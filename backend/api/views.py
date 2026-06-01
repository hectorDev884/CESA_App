from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from api.models import Estudiante, Beca, AsistenciaBeca
from api.serializers import EstudianteSerializer, BecaSerializer, AsistenciaBecaSerializer
from rest_framework.filters import SearchFilter, OrderingFilter
from django.db.models import Count, Q, F
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib.units import mm
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

WEEKDAY_NAMES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

def generar_pdf_asistencia(request):
    """
    GET /api/pdf/asistencia/?nc=22290697&nombre=Alan%20Emiliano&fecha_inicio=2025-11-01&fecha_fin=2025-11-30&color=green
    """
    nc = request.GET.get("nc")
    nombre = request.GET.get("nombre", "")
    start = request.GET.get("fecha_inicio", datetime.today())
    end = request.GET.get("fecha_fin", "2025-11-30")
    color_param = request.GET.get("color", "red")

    if not (nc and start and end):
        return HttpResponseBadRequest("Faltan parámetros: nc, start, end")

    try:
        fecha_inicio = datetime.strptime(start, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(end, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponseBadRequest("Fechas deben tener formato YYYY-MM-DD")

    if fecha_fin < fecha_inicio:
        return HttpResponseBadRequest("La fecha fin debe ser posterior o igual a la fecha inicio")

# --- Colores predefinidos ---
    COLOR_MAP = {
        "red": (0.75, 0.12, 0.15),
        "green": (0.0, 0.6, 0.3),
        "blue": (0.0, 0.3, 0.7),
        "orange": (1.0, 0.5, 0.0),
        "purple": (0.5, 0.0, 0.5),
        "teal": (0.0, 0.5, 0.5),
        "yellow": (1.0, 0.9, 0.0),
        "pink": (1.0, 0.4, 0.6),
        "gray": (0.5, 0.5, 0.5),
        "brown": (0.6, 0.4, 0.2),
    }

    header_color = COLOR_MAP.get(color_param.lower(), (0.75, 0.12, 0.15))

    # --- Calcular primer lunes ---
    # Si fecha_inicio cae en sabado (5), no retrocedemos
    if fecha_inicio.weekday() == 5:  # sabado
        primer_lunes = fecha_inicio + timedelta(days=2)
    elif fecha_inicio.weekday() == 6:
        primer_lunes = fecha_inicio + timedelta(days=1)
    else:
        primer_lunes = fecha_inicio - timedelta(days=fecha_inicio.weekday())

    semanas = []
    cursor = primer_lunes
    while cursor <= fecha_fin:
        semana = []
        for d in range(5):
            dia_fecha = cursor + timedelta(days=d)
            if fecha_inicio <= dia_fecha <= fecha_fin:
                fecha_str = dia_fecha.strftime("%d/%m/%Y")
                semana.append({"dia": WEEKDAY_NAMES[d], "fecha": fecha_str})
            else:
                semana.append({"dia": WEEKDAY_NAMES[d], "fecha": ""})
        semanas.append(semana)
        cursor += timedelta(weeks=1)

    # --- Crear PDF ---
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="asistencia_{nc}.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Márgenes
    left_margin = 20 * mm
    right_margin = 20 * mm
    top_margin = 20 * mm
    usable_width = width - left_margin - right_margin
    header_height = 18 * mm

    # Dibujar encabezado
    pdf.setFillColorRGB(*header_color)
    pdf.rect(left_margin, height - top_margin - header_height, usable_width, header_height, stroke=0, fill=1)

    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawCentredString(left_margin + usable_width / 2, height - top_margin - header_height/2 + 4, "FORMATO DE ASISTENCIA")

    # Datos estudiante
    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica", 11)
    y = height - top_margin - header_height - 12
    pdf.drawString(left_margin, y, f"NC: {nc}")
    pdf.drawString(left_margin + 220, y, f"Nombre completo: {nombre}")
    y -= 16
    pdf.drawString(left_margin, y, f"Período: {fecha_inicio.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}")
    y -= 22

    # --- Construir tabla ---
    col_width = usable_width / 5.0
    data = []
    for semana in semanas:
        fila_fechas = []
        for cel in semana:
            texto = cel["dia"]
            if cel["fecha"]:
                texto = f"{texto}\n{cel['fecha']}"
            fila_fechas.append(texto)
        data.append(fila_fechas)
        data.append([""] * 5)

    if not data:
        data = [[f"{d}\n" for d in WEEKDAY_NAMES], [""] * 5]

    fila_alto_fecha = 12 * mm
    fila_alto_firma = 20 * mm
    row_heights = [fila_alto_fecha if i % 2 == 0 else fila_alto_firma for i in range(len(data))]

    style = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ])

    # --- Dibujar tabla con paginado ---
    pdf_y_start = y
    current_row = 0
    total_rows = len(data)

    while current_row < total_rows:
        remaining_height = pdf_y_start - 40
        h_sum = 0
        rows_fit = 0
        for r in range(current_row, total_rows):
            h_sum += row_heights[r]
            if h_sum <= remaining_height:
                rows_fit += 1
            else:
                break
        if rows_fit == 0:
            rows_fit = 1

        sub_data = data[current_row: current_row + rows_fit]
        sub_heights = row_heights[current_row: current_row + rows_fit]
        sub_table = Table(sub_data, colWidths=[col_width]*5, rowHeights=sub_heights)
        sub_table.setStyle(style)
        sub_table.wrapOn(pdf, left_margin, pdf_y_start)
        sub_table.drawOn(pdf, left_margin, pdf_y_start - sum(sub_heights))

        pdf_y_start -= sum(sub_heights) + 10
        current_row += rows_fit

        if current_row < total_rows:
            pdf.showPage()
            pdf.setFillColorRGB(*header_color)
            pdf.rect(left_margin, height - top_margin - header_height, usable_width, header_height, stroke=0, fill=1)
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 14)
            pdf.drawCentredString(left_margin + usable_width / 2, height - top_margin - header_height/2 + 4, "FORMATO DE ASISTENCIA")
            pdf.setFillColor(colors.black)
            pdf.setFont("Helvetica", 11)
            y_page = height - top_margin - header_height - 12
            pdf.drawString(left_margin, y_page, f"NC: {nc}")
            pdf.drawString(left_margin + 220, y_page, f"Nombre completo: {nombre}")
            y_page -= 16
            pdf.drawString(left_margin, y_page, f"Período: {fecha_inicio.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}")
            pdf_y_start = y_page - 22

    # Firma
    pdf.setFont("Helvetica", 11)
    

    pdf.showPage()
    pdf.save()
    return response


def generar_pdf_asistencia_general(request):
    """
    Genera un solo PDF con todos los alumnos con beca aprobada.
    Cada alumno tiene su formato de asistencia en páginas consecutivas.
    Cada 10 alumnos cambia el color de encabezado.
    """
    start = request.GET.get("fecha_inicio", datetime.today().strftime("%Y-%m-%d"))
    end = request.GET.get("fecha_fin", "2025-11-30")

    try:
        fecha_inicio = datetime.strptime(start, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(end, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponseBadRequest("Fechas deben tener formato YYYY-MM-DD")

    if fecha_fin < fecha_inicio:
        return HttpResponseBadRequest("La fecha fin debe ser posterior o igual a la fecha inicio")

    COLOR_MAP = {
    "yellow": (0.894, 0.922, 0.157),
    "red": (0.722, 0.200, 0.086),
    "green": (0.310, 0.722, 0.086),
    "blue": (0.086, 0.565, 0.722),
    "pink": (0.800, 0.094, 0.753),
}

    color_names = list(COLOR_MAP.keys())

    # --- Obtener becas aprobadas ---
    becas = Beca.objects.filter(estatus="aprobada").select_related("numero_control")
    if not becas.exists():
        return HttpResponseBadRequest("No hay becas aprobadas en este momento")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="asistencia_general.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Márgenes
    left_margin = 20 * mm
    right_margin = 20 * mm
    top_margin = 20 * mm
    usable_width = width - left_margin - right_margin
    header_height = 18 * mm

    # --- Calcular primer lunes ---
    def get_weeks(fecha_inicio, fecha_fin):
        if fecha_inicio.weekday() == 5:
            primer_lunes = fecha_inicio + timedelta(days=2)
        elif fecha_inicio.weekday() == 6:
            primer_lunes = fecha_inicio + timedelta(days=1)
        else:
            primer_lunes = fecha_inicio - timedelta(days=fecha_inicio.weekday())

        semanas = []
        cursor = primer_lunes
        while cursor <= fecha_fin:
            semana = []
            for d in range(5):
                dia_fecha = cursor + timedelta(days=d)
                if fecha_inicio <= dia_fecha <= fecha_fin:
                    fecha_str = dia_fecha.strftime("%d/%m/%Y")
                    semana.append({"dia": WEEKDAY_NAMES[d], "fecha": fecha_str})
                else:
                    semana.append({"dia": WEEKDAY_NAMES[d], "fecha": ""})
            semanas.append(semana)
            cursor += timedelta(weeks=1)
        return semanas

    # --- Generar una hoja por alumno ---
    for i, beca in enumerate(becas):
        estudiante = beca.numero_control
        nombre = f"{estudiante.nombre} {estudiante.apellido}"
        nc = estudiante.numero_control

        # Cambiar color cada 15 alumnos
        color_index = (i // 15) % len(color_names)
        header_color = COLOR_MAP[color_names[color_index]]

        # Dibujar encabezado
        pdf.setFillColorRGB(*header_color)
        pdf.rect(left_margin, height - top_margin - header_height, usable_width, header_height, stroke=0, fill=1)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawCentredString(left_margin + usable_width / 2, height - top_margin - header_height / 2 + 4, "FORMATO DE ASISTENCIA")

        # Datos alumno
        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica", 11)
        y = height - top_margin - header_height - 12
        pdf.drawString(left_margin, y, f"NC: {nc}")
        pdf.drawString(left_margin + 220, y, f"Nombre completo: {nombre}")
        y -= 16
        pdf.drawString(left_margin, y, f"Período: {fecha_inicio.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}")
        y -= 22

        # --- Tabla de asistencia ---
        semanas = get_weeks(fecha_inicio, fecha_fin)
        col_width = usable_width / 5.0
        data = []
        for semana in semanas:
            fila_fechas = []
            for cel in semana:
                texto = cel["dia"]
                if cel["fecha"]:
                    texto = f"{texto}\n{cel['fecha']}"
                fila_fechas.append(texto)
            data.append(fila_fechas)
            data.append([""] * 5)

        fila_alto_fecha = 12 * mm
        fila_alto_firma = 20 * mm
        row_heights = [fila_alto_fecha if i % 2 == 0 else fila_alto_firma for i in range(len(data))]

        style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ])

        table = Table(data, colWidths=[col_width]*5, rowHeights=row_heights)
        table.setStyle(style)
        table.wrapOn(pdf, left_margin, y)
        table.drawOn(pdf, left_margin, y - sum(row_heights))

        # Salto de página para el siguiente alumno
        pdf.showPage()

    pdf.save()
    return response


from rest_framework.pagination import PageNumberPagination

class EstudiantePagination(PageNumberPagination):
    page_size = 10  # máximo 10 registros por página

class EstudianteViewSet(viewsets.ModelViewSet):
    queryset = Estudiante.objects.all()
    serializer_class = EstudianteSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['numero_control', 'nombre', 'apellido', 'email']
    ordering_fields = ['numero_control', 'apellido', 'nombre', 'fecha_registro']
    ordering = ['apellido', 'nombre']
    pagination_class = EstudiantePagination

    @action(detail=False, methods=['post'], url_path='importar')
    def importar_excel(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'No se proporcionó archivo'}, status=400)

        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Archivo inválido: {str(e)}'}, status=400)

        importados = 0
        errores = []

        for fila_num in range(2, ws.max_row + 1):
            vals = [ws.cell(row=fila_num, column=c).value for c in range(1, 8)]

            if all(v is None or str(v).strip() == '' for v in vals):
                continue

            nc       = str(vals[0]).strip() if vals[0] is not None else ''
            nombre   = str(vals[1]).strip() if vals[1] is not None else ''
            apellido = str(vals[2]).strip() if vals[2] is not None else ''
            email    = str(vals[3]).strip() if vals[3] is not None else ''
            carrera  = str(vals[4]).strip() if vals[4] is not None else ''
            telefono = str(vals[6]).strip() if vals[6] is not None else ''

            if not nc:
                errores.append({'fila': fila_num, 'error': 'Número de control vacío'})
                continue
            if not nombre:
                errores.append({'fila': fila_num, 'error': f'NC {nc}: Nombre vacío'})
                continue
            if not apellido:
                errores.append({'fila': fila_num, 'error': f'NC {nc}: Apellido vacío'})
                continue
            if not email or '@' not in email:
                errores.append({'fila': fila_num, 'error': f'NC {nc}: Email inválido'})
                continue

            semestre = None
            if vals[5] is not None:
                try:
                    semestre = int(vals[5])
                except (ValueError, TypeError):
                    errores.append({'fila': fila_num, 'error': f'NC {nc}: Semestre debe ser número entero'})
                    continue

            try:
                Estudiante.objects.update_or_create(
                    numero_control=nc,
                    defaults={
                        'nombre': nombre,
                        'apellido': apellido,
                        'email': email,
                        'carrera': carrera,
                        'semestre': semestre,
                        'telefono': telefono,
                    }
                )
                importados += 1
            except Exception as e:
                errores.append({'fila': fila_num, 'error': f'NC {nc}: {str(e)}'})

        return Response({
            'importados': importados,
            'errores': errores,
            'total_errores': len(errores),
        })


class BecaPagination(PageNumberPagination):
    page_size = 10

class BecaViewSet(viewsets.ModelViewSet):
    queryset = Beca.objects.all()
    serializer_class = BecaSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['tipo_beca', 'estatus', 'numero_control__numero_control',
                     'numero_control__nombre', 'numero_control__apellido']
    ordering_fields = ['tipo_beca', 'estatus']
    pagination_class = EstudiantePagination

class AsistenciaBecaViewSet(viewsets.ModelViewSet):
    queryset = AsistenciaBeca.objects.all()
    serializer_class = AsistenciaBecaSerializer
    filter_backends = [SearchFilter]
    search_fields = ['beca_id__beca_id', 'asistencia_id']


def reporte_becas(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    carrera_filtro = request.GET.get('carrera')

    qs = Beca.objects.select_related('numero_control').all()

    if fecha_inicio:
        try:
            fi = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            qs = qs.filter(fecha_solicitud__gte=fi)
        except ValueError:
            pass
    if fecha_fin:
        try:
            ff = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            qs = qs.filter(fecha_solicitud__lte=ff)
        except ValueError:
            pass
    if carrera_filtro:
        qs = qs.filter(numero_control__carrera=carrera_filtro)

    total     = qs.count()
    aprobadas = qs.filter(estatus='aprobada').count()
    pendientes = qs.filter(estatus='pendiente').count()

    por_carrera = list(
        qs.values(carrera=F('numero_control__carrera'))
        .annotate(
            total=Count('beca_id'),
            aprobadas=Count('beca_id', filter=Q(estatus='aprobada')),
            pendientes=Count('beca_id', filter=Q(estatus='pendiente')),
        )
        .order_by('-total')
    )

    por_tipo = list(
        qs.values('tipo_beca')
        .annotate(total=Count('beca_id'))
        .order_by('-total')
    )

    por_semestre = list(
        qs.values(semestre=F('numero_control__semestre'))
        .annotate(total=Count('beca_id'))
        .filter(semestre__isnull=False)
        .order_by('-total')
    )

    por_estatus = list(
        qs.values('estatus')
        .annotate(total=Count('beca_id'))
        .order_by('-total')
    )

    tipo_x_carrera = list(
        qs.values(
            tipo=F('tipo_beca'),
            carrera=F('numero_control__carrera'),
        )
        .annotate(total=Count('beca_id'))
        .order_by('tipo', 'carrera')
    )

    por_mes_solicitud = list(
        qs.values(mes=F('fecha_solicitud__month'), anio=F('fecha_solicitud__year'))
        .annotate(total=Count('beca_id'))
        .filter(fecha_solicitud__isnull=False)
        .order_by('-anio', '-mes')
    )

    hoy = date.today()
    anio = hoy.year
    sem_actual = 'A' if hoy.month <= 6 else 'B'

    semestres_academicos = []
    sem, anio_cur = sem_actual, anio
    for _ in range(6):
        if sem == 'A':
            inicio = date(anio_cur, 1, 1)
            fin    = date(anio_cur, 6, 30)
        else:
            inicio = date(anio_cur, 7, 1)
            fin    = date(anio_cur, 12, 31)
        total_sem = qs.filter(
            fecha_solicitud__gte=inicio,
            fecha_solicitud__lte=fin
        ).count()
        semestres_academicos.append({
            'semestre': f'{anio_cur}-{sem}',
            'total': total_sem,
        })
        if sem == 'A':
            sem = 'B'
            anio_cur -= 1
        else:
            sem = 'A'

    semestres_academicos.reverse()

    return JsonResponse({
        'totales': {'total': total, 'aprobadas': aprobadas, 'pendientes': pendientes},
        'por_carrera': por_carrera,
        'por_tipo': por_tipo,
        'por_semestre': por_semestre,
        'por_estatus': por_estatus,
        'tipo_x_carrera': tipo_x_carrera,
        'por_mes_solicitud': por_mes_solicitud,
        'por_semestre_academico': semestres_academicos,
    })


def exportar_becas_excel(request):
    wb = openpyxl.Workbook()
    fill_verde = PatternFill("solid", fgColor="036942")
    font_blanco = Font(bold=True, color="FFFFFF")
    centrado = Alignment(horizontal='center', vertical='center', wrap_text=True)
    borde_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font_blanco
            cell.fill = fill_verde
            cell.alignment = centrado
            cell.border = borde_thin

    def _format_sheet(ws, col_widths=None):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.row > 1:
                    cell.border = borde_thin
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        else:
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    val = str(cell.value) if cell.value is not None else ''
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.tabColor = "036942"

    # Hoja 1: todas las becas
    ws1 = wb.active
    ws1.title = "Becas"
    _header(ws1, ['ID', 'Número Control', 'Estudiante', 'Carrera', 'Tipo Beca',
                  'Fecha Solicitud', 'Fecha Aprobación', 'Fecha Entrega', 'Fecha Fin', 'Estatus'])
    for row, b in enumerate(Beca.objects.select_related('numero_control').order_by('beca_id'), 2):
        e = b.numero_control
        ws1.cell(row=row, column=1, value=b.beca_id)
        ws1.cell(row=row, column=2, value=e.numero_control if e else '')
        ws1.cell(row=row, column=3, value=f"{e.nombre} {e.apellido}" if e else '')
        ws1.cell(row=row, column=4, value=e.carrera if e else '')
        ws1.cell(row=row, column=5, value=b.tipo_beca)
        ws1.cell(row=row, column=6, value=str(b.fecha_solicitud) if b.fecha_solicitud else '')
        ws1.cell(row=row, column=7, value=str(b.fecha_aprobacion) if b.fecha_aprobacion else '')
        ws1.cell(row=row, column=8, value=str(b.fecha_entrega) if b.fecha_entrega else '')
        ws1.cell(row=row, column=9, value=str(b.fecha_fin) if b.fecha_fin else '')
        ws1.cell(row=row, column=10, value=b.estatus)
    _format_sheet(ws1)

    # Hoja 2: por carrera
    ws2 = wb.create_sheet("Por Carrera")
    _header(ws2, ['Carrera', 'Total', 'Aprobadas', 'Pendientes'])
    for row, item in enumerate(
        Beca.objects
        .values(carrera=F('numero_control__carrera'))
        .annotate(total=Count('beca_id'),
                  aprobadas=Count('beca_id', filter=Q(estatus='aprobada')),
                  pendientes=Count('beca_id', filter=Q(estatus='pendiente')))
        .order_by('-total'),
        2
    ):
        ws2.cell(row=row, column=1, value=item['carrera'] or 'Sin carrera')
        ws2.cell(row=row, column=2, value=item['total'])
        ws2.cell(row=row, column=3, value=item['aprobadas'])
        ws2.cell(row=row, column=4, value=item['pendientes'])
    _format_sheet(ws2)

    # Hoja 3: por tipo
    ws3 = wb.create_sheet("Por Tipo")
    _header(ws3, ['Tipo de Beca', 'Total'])
    for row, item in enumerate(
        Beca.objects.values('tipo_beca').annotate(total=Count('beca_id')).order_by('-total'),
        2
    ):
        ws3.cell(row=row, column=1, value=item['tipo_beca'] or 'Sin tipo')
        ws3.cell(row=row, column=2, value=item['total'])
    _format_sheet(ws3)

    # Hoja 4: por estatus
    ws4 = wb.create_sheet("Por Estatus")
    _header(ws4, ['Estatus', 'Total'])
    for row, item in enumerate(
        Beca.objects.values('estatus').annotate(total=Count('beca_id')).order_by('-total'),
        2
    ):
        ws4.cell(row=row, column=1, value=item['estatus'] or 'Sin estatus')
        ws4.cell(row=row, column=2, value=item['total'])
    _format_sheet(ws4)

    # Hoja 5: Tipo x Carrera (tabla pivote)
    ws5 = wb.create_sheet("Tipo x Carrera")
    _header(ws5, ['Tipo de Beca', 'Carrera', 'Total'])
    for row, item in enumerate(
        Beca.objects
        .values(tipo=F('tipo_beca'), carrera=F('numero_control__carrera'))
        .annotate(total=Count('beca_id'))
        .order_by('tipo', '-total'),
        2
    ):
        ws5.cell(row=row, column=1, value=item['tipo'] or 'Sin tipo')
        ws5.cell(row=row, column=2, value=item['carrera'] or 'Sin carrera')
        ws5.cell(row=row, column=3, value=item['total'])
    _format_sheet(ws5)

    # Hoja 6: por mes de solicitud
    ws6 = wb.create_sheet("Por Mes")
    _header(ws6, ['Año', 'Mes', 'Total'])
    for row, item in enumerate(
        Beca.objects
        .filter(fecha_solicitud__isnull=False)
        .values(anio=F('fecha_solicitud__year'), mes=F('fecha_solicitud__month'))
        .annotate(total=Count('beca_id'))
        .order_by('-anio', '-mes'),
        2
    ):
        ws6.cell(row=row, column=1, value=item['anio'])
        ws6.cell(row=row, column=2, value=item['mes'])
        ws6.cell(row=row, column=3, value=item['total'])
    _format_sheet(ws6)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="reporte_becas.xlsx"'
    return resp


def reporte_estudiantes(request):
    qs = Estudiante.objects.all()

    total = qs.count()

    por_carrera = list(
        qs.values('carrera')
        .annotate(total=Count('numero_control'))
        .order_by('-total')
    )

    por_semestre = list(
        qs.values('semestre')
        .annotate(total=Count('numero_control'))
        .filter(semestre__isnull=False)
        .order_by('semestre')
    )

    carrera_x_semestre = list(
        qs.values('carrera', 'semestre')
        .annotate(total=Count('numero_control'))
        .filter(carrera__isnull=False, semestre__isnull=False)
        .order_by('carrera', 'semestre')
    )

    por_mes_registro = list(
        qs.filter(fecha_registro__isnull=False)
        .values(mes=F('fecha_registro__month'), anio=F('fecha_registro__year'))
        .annotate(total=Count('numero_control'))
        .order_by('-anio', '-mes')[:12]
    )

    return JsonResponse({
        'totales': {'total': total},
        'por_carrera': por_carrera,
        'por_semestre': por_semestre,
        'carrera_x_semestre': carrera_x_semestre,
        'por_mes_registro': list(reversed(por_mes_registro)),
    })


def exportar_estudiantes_excel(request):
    wb = openpyxl.Workbook()
    fill_verde = PatternFill("solid", fgColor="036942")
    font_blanco = Font(bold=True, color="FFFFFF")
    centrado = Alignment(horizontal='center', vertical='center', wrap_text=True)
    borde_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font_blanco
            cell.fill = fill_verde
            cell.alignment = centrado
            cell.border = borde_thin

    def _format_sheet(ws):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.row > 1:
                    cell.border = borde_thin
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 55)
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.tabColor = "036942"

    ws1 = wb.active
    ws1.title = "Estudiantes"
    _header(ws1, ['Número Control', 'Nombre', 'Apellido', 'Email', 'Carrera',
                  'Semestre', 'Teléfono', 'Fecha Registro'])
    for row, e in enumerate(Estudiante.objects.order_by('apellido', 'nombre'), 2):
        ws1.cell(row=row, column=1, value=e.numero_control)
        ws1.cell(row=row, column=2, value=e.nombre)
        ws1.cell(row=row, column=3, value=e.apellido)
        ws1.cell(row=row, column=4, value=e.email)
        ws1.cell(row=row, column=5, value=e.carrera or '')
        ws1.cell(row=row, column=6, value=e.semestre or '')
        ws1.cell(row=row, column=7, value=e.telefono or '')
        ws1.cell(row=row, column=8, value=str(e.fecha_registro)[:10] if e.fecha_registro else '')
    _format_sheet(ws1)

    ws2 = wb.create_sheet("Por Carrera")
    _header(ws2, ['Carrera', 'Total'])
    for row, item in enumerate(
        Estudiante.objects.values('carrera')
        .annotate(total=Count('numero_control'))
        .order_by('-total'),
        2
    ):
        ws2.cell(row=row, column=1, value=item['carrera'] or 'Sin carrera')
        ws2.cell(row=row, column=2, value=item['total'])
    _format_sheet(ws2)

    ws3 = wb.create_sheet("Por Semestre")
    _header(ws3, ['Semestre', 'Total'])
    for row, item in enumerate(
        Estudiante.objects.values('semestre')
        .annotate(total=Count('numero_control'))
        .filter(semestre__isnull=False)
        .order_by('semestre'),
        2
    ):
        ws3.cell(row=row, column=1, value=item['semestre'])
        ws3.cell(row=row, column=2, value=item['total'])
    _format_sheet(ws3)

    ws4 = wb.create_sheet("Carrera x Semestre")
    _header(ws4, ['Carrera', 'Semestre', 'Total'])
    for row, item in enumerate(
        Estudiante.objects.values('carrera', 'semestre')
        .annotate(total=Count('numero_control'))
        .filter(carrera__isnull=False, semestre__isnull=False)
        .order_by('carrera', 'semestre'),
        2
    ):
        ws4.cell(row=row, column=1, value=item['carrera'])
        ws4.cell(row=row, column=2, value=item['semestre'])
        ws4.cell(row=row, column=3, value=item['total'])
    _format_sheet(ws4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="reporte_estudiantes.xlsx"'
    return resp

